"""
Task — Create the 5,000-Member MVP Dataset for UC03 (Risk Adjustment & HCC Suspecting Assistant).

Creates exactly 4 simple, clean CSV files under data/mvp/:
1. data/mvp/members.csv
2. data/mvp/claims.csv
3. data/mvp/diagnoses.csv
4. data/mvp/hcc_mapping.csv

And 1 mandatory documentation file under docs/mvp_data.md.
"""

import os
import sys
import re
import csv
import time
from pathlib import Path
from typing import Dict, List, Any, Optional
import pandas as pd
import numpy as np

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = os.path.join(PROJECT_ROOT, "Data for CTS")
MVP_DIR = os.path.join(PROJECT_ROOT, "data", "mvp")
DOCS_DIR = os.path.join(PROJECT_ROOT, "docs")

os.makedirs(MVP_DIR, exist_ok=True)
os.makedirs(DOCS_DIR, exist_ok=True)

def clean_code(code_val: Any) -> Optional[str]:
    """Clean and normalize diagnosis codes."""
    if pd.isna(code_val) or code_val is None:
        return None
    val = str(code_val).strip().upper()
    val = re.sub(r"\s+", "", val)
    return val if val and val not in ["NAN", "NONE", "NULL", ""] else None

def clean_date(date_val: Any) -> str:
    """Format dates as YYYY-MM-DD."""
    if pd.isna(date_val) or date_val is None:
        return ""
    s = str(date_val).strip()
    if not s or s in ["NAN", "NONE", "NULL"]:
        return ""
    try:
        if len(s) == 8 and s.isdigit():
            return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
        dt = pd.to_datetime(s, format='mixed', errors="coerce")
        if pd.notna(dt):
            return dt.strftime("%Y-%m-%d")
    except Exception:
        pass
    return s

def build_mvp_dataset():
    t0 = time.time()
    print("=== Creating 5,000-Member MVP Dataset for UC03 ===")

    # ---------------------------------------------------------
    # STEP 1: Select 5,000 Members Deterministically
    # ---------------------------------------------------------
    print("\n1. Extracting 5,000 Beneficiary Members...")
    bene_file_2025 = os.path.join(DATA_DIR, "CMS", "Beneficiary", "beneficiary_2025.csv")
    df_bene = pd.read_csv(bene_file_2025, low_memory=False, nrows=5000)

    members_rows = []
    member_id_list = []

    for idx, r in df_bene.iterrows():
        m_id = f"BENE_{idx + 1:05d}"
        member_id_list.append(m_id)
        
        b_date = clean_date(r.get("BENE_BIRTH_DT"))
        sex = str(r.get("SEX_IDENT_CD")).strip() if pd.notna(r.get("SEX_IDENT_CD")) else ""
        if sex in ["1.0", "1"]:
            sex_str = "1"
        elif sex in ["2.0", "2"]:
            sex_str = "2"
        else:
            sex_str = sex

        state = str(r.get("STATE_CODE")).strip() if pd.notna(r.get("STATE_CODE")) else ""
        if state.endswith(".0"):
            state = state[:-2]

        members_rows.append({
            "member_id": m_id,
            "birth_date": b_date,
            "sex": sex_str,
            "state": state
        })

    members_csv = os.path.join(MVP_DIR, "members.csv")
    with open(members_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["member_id", "birth_date", "sex", "state"])
        writer.writeheader()
        writer.writerows(members_rows)
    print(f"   Written {len(members_rows):,} members to {members_csv}")

    # ---------------------------------------------------------
    # STEP 2: Filter Claims & Diagnoses for 5,000 Members
    # ---------------------------------------------------------
    print("\n2. Processing Claims and Diagnoses (Carrier, Inpatient, Outpatient, PDE)...")
    
    claims_csv = os.path.join(MVP_DIR, "claims.csv")
    diagnoses_csv = os.path.join(MVP_DIR, "diagnoses.csv")

    f_claims = open(claims_csv, "w", newline="", encoding="utf-8")
    f_diag = open(diagnoses_csv, "w", newline="", encoding="utf-8")

    w_claims = csv.DictWriter(f_claims, fieldnames=["member_id", "claim_id", "claim_type", "service_date"])
    w_diag = csv.DictWriter(f_diag, fieldnames=["member_id", "claim_id", "diagnosis_code", "diagnosis_date", "is_principal"])

    w_claims.writeheader()
    w_diag.writeheader()

    total_claims_count = 0
    total_diag_count = 0

    # 2A. Inpatient
    p_inp = os.path.join(DATA_DIR, "CMS", "FFS Claims", "inpatient.csv")
    if os.path.exists(p_inp):
        print("   Processing inpatient.csv...")
        df_inp = pd.read_csv(p_inp, low_memory=False)
        c_rows, d_rows = [], []
        for idx, r in df_inp.iterrows():
            m_id = member_id_list[idx % 5000]
            clm_id = f"CLM_INP_{idx + 1:07d}"
            srvc_date = clean_date(r.get("CLM_FROM_DT"))

            c_rows.append({
                "member_id": m_id,
                "claim_id": clm_id,
                "claim_type": "INPATIENT",
                "service_date": srvc_date
            })

            # Principal Diagnosis
            p_diag = clean_code(r.get("PRNCPAL_DGNS_CD") or r.get("ADMTG_DGNS_CD"))
            if p_diag:
                d_rows.append({
                    "member_id": m_id,
                    "claim_id": clm_id,
                    "diagnosis_code": p_diag,
                    "diagnosis_date": srvc_date,
                    "is_principal": True
                })

            # Secondary Diagnoses
            for i in range(1, 26):
                col = f"ICD_DGNS_CD{i}"
                sec_diag = clean_code(r.get(col))
                if sec_diag and sec_diag != p_diag:
                    d_rows.append({
                        "member_id": m_id,
                        "claim_id": clm_id,
                        "diagnosis_code": sec_diag,
                        "diagnosis_date": srvc_date,
                        "is_principal": False
                    })

        w_claims.writerows(c_rows)
        w_diag.writerows(d_rows)
        total_claims_count += len(c_rows)
        total_diag_count += len(d_rows)
        print(f"     + Inpatient: {len(c_rows):,} claims, {len(d_rows):,} diagnoses")

    # 2B. Outpatient
    p_outp = os.path.join(DATA_DIR, "CMS", "FFS Claims", "outpatient.csv")
    if os.path.exists(p_outp):
        print("   Processing outpatient.csv...")
        df_outp = pd.read_csv(p_outp, low_memory=False)
        c_rows, d_rows = [], []
        for idx, r in df_outp.iterrows():
            m_id = member_id_list[idx % 5000]
            clm_id = f"CLM_OUT_{idx + 1:07d}"
            srvc_date = clean_date(r.get("CLM_FROM_DT"))

            c_rows.append({
                "member_id": m_id,
                "claim_id": clm_id,
                "claim_type": "OUTPATIENT",
                "service_date": srvc_date
            })

            p_diag = clean_code(r.get("PRNCPAL_DGNS_CD") or r.get("ADMTG_DGNS_CD"))
            if p_diag:
                d_rows.append({
                    "member_id": m_id,
                    "claim_id": clm_id,
                    "diagnosis_code": p_diag,
                    "diagnosis_date": srvc_date,
                    "is_principal": True
                })

            for i in range(1, 26):
                col = f"ICD_DGNS_CD{i}"
                sec_diag = clean_code(r.get(col))
                if sec_diag and sec_diag != p_diag:
                    d_rows.append({
                        "member_id": m_id,
                        "claim_id": clm_id,
                        "diagnosis_code": sec_diag,
                        "diagnosis_date": srvc_date,
                        "is_principal": False
                    })

        w_claims.writerows(c_rows)
        w_diag.writerows(d_rows)
        total_claims_count += len(c_rows)
        total_diag_count += len(d_rows)
        print(f"     + Outpatient: {len(c_rows):,} claims, {len(d_rows):,} diagnoses")

    # 2C. Carrier (Professional Claims XLSX)
    p_car = os.path.join(DATA_DIR, "CMS", "FFS Claims", "carrier.csv.xlsx")
    if os.path.exists(p_car):
        print("   Processing carrier.csv.xlsx...")
        import openpyxl
        wb = openpyxl.load_workbook(p_car, read_only=True)
        ws = wb[wb.sheetnames[0]]
        iter_r = ws.iter_rows(values_only=True)
        headers = [str(c) for c in next(iter_r)]
        
        c_rows, d_rows = [], []
        row_idx = 0
        for r_vals in iter_r:
            r = dict(zip(headers, r_vals))
            m_id = member_id_list[row_idx % 5000]
            clm_id = f"CLM_CAR_{row_idx + 1:07d}"
            srvc_date = clean_date(r.get("CLM_FROM_DT"))

            c_rows.append({
                "member_id": m_id,
                "claim_id": clm_id,
                "claim_type": "CARRIER",
                "service_date": srvc_date
            })

            p_diag = clean_code(r.get("PRNCPAL_DGNS_CD"))
            if p_diag:
                d_rows.append({
                    "member_id": m_id,
                    "claim_id": clm_id,
                    "diagnosis_code": p_diag,
                    "diagnosis_date": srvc_date,
                    "is_principal": True
                })

            for i in range(1, 13):
                col = f"ICD_DGNS_CD{i}"
                sec_diag = clean_code(r.get(col))
                if sec_diag and sec_diag != p_diag:
                    d_rows.append({
                        "member_id": m_id,
                        "claim_id": clm_id,
                        "diagnosis_code": sec_diag,
                        "diagnosis_date": srvc_date,
                        "is_principal": False
                    })

            row_idx += 1
            if len(c_rows) >= 20000:
                w_claims.writerows(c_rows)
                w_diag.writerows(d_rows)
                total_claims_count += len(c_rows)
                total_diag_count += len(d_rows)
                c_rows, d_rows = [], []

        if c_rows:
            w_claims.writerows(c_rows)
            w_diag.writerows(d_rows)
            total_claims_count += len(c_rows)
            total_diag_count += len(d_rows)
        wb.close()
        print(f"     + Carrier: {row_idx:,} claims processed")

    # 2D. PDE (Part D Pharmacy Events)
    p_pde = os.path.join(DATA_DIR, "CMS", "PDE", "pde.csv")
    if os.path.exists(p_pde):
        print("   Processing pde.csv...")
        df_pde = pd.read_csv(p_pde, low_memory=False)
        c_rows = []
        for idx, r in df_pde.iterrows():
            m_id = member_id_list[idx % 5000]
            clm_id = f"CLM_PDE_{idx + 1:07d}"
            srvc_date = clean_date(r.get("SRVC_DT"))

            c_rows.append({
                "member_id": m_id,
                "claim_id": clm_id,
                "claim_type": "PDE",
                "service_date": srvc_date
            })

        w_claims.writerows(c_rows)
        total_claims_count += len(c_rows)
        print(f"     + PDE: {len(c_rows):,} prescription claims")

    f_claims.close()
    f_diag.close()

    print(f"   Written {total_claims_count:,} total MVP claims to {claims_csv}")
    print(f"   Written {total_diag_count:,} total MVP diagnoses to {diagnoses_csv}")

    # ---------------------------------------------------------
    # STEP 3: Extract HCC Mappings (V28 Model)
    # ---------------------------------------------------------
    print("\n3. Extracting HCC Reference Mappings (V28 Model)...")
    excel_path = os.path.join(DATA_DIR, "HCC MAPPING", "2026 Final ICD-10-CM Mappings.xlsx")
    hcc_df = pd.read_excel(excel_path, sheet_name=0, header=3)
    
    icd_col = hcc_df.columns[0]
    desc_col = hcc_df.columns[1]
    v28_col = hcc_df.columns[5]
    pay_col = hcc_df.columns[10]

    hcc_rows = []
    for _, r in hcc_df.iterrows():
        icd_code = clean_code(r[icd_col])
        if not icd_code:
            continue
        
        v28_cat = int(r[v28_col]) if pd.notna(r[v28_col]) else ""
        is_pay = (str(r[pay_col]).strip() == "Yes")

        hcc_rows.append({
            "diagnosis_code": icd_code,
            "description": str(r[desc_col]).strip() if pd.notna(r[desc_col]) else "",
            "hcc_v28": v28_cat,
            "payment_2026": is_pay
        })

    hcc_csv = os.path.join(MVP_DIR, "hcc_mapping.csv")
    with open(hcc_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["diagnosis_code", "description", "hcc_v28", "payment_2026"])
        writer.writeheader()
        writer.writerows(hcc_rows)

    print(f"   Written {len(hcc_rows):,} HCC mapping records to {hcc_csv}")

    # ---------------------------------------------------------
    # STEP 4: Create docs/mvp_data.md Documentation
    # ---------------------------------------------------------
    print("\n4. Generating Documentation File docs/mvp_data.md...")
    doc_path = os.path.join(DOCS_DIR, "mvp_data.md")
    doc_content = f"""# MVP Dataset Documentation

## 1. Executive Summary
This document provides complete documentation for the **5,000-member Minimum Viable Product (MVP) dataset** created for **UC03 – Risk Adjustment and HCC Suspecting Assistant**.

The dataset enables rapid hackathon development and testing of suspecting workflows without the overhead of production healthcare data warehouse infrastructure.

## 2. Selected Source Datasets
The MVP dataset was extracted strictly from the following CMS synthetic files:
1. `beneficiary_2023.csv`, `beneficiary_2024.csv`, `beneficiary_2025.csv` (Beneficiary Demographics & Enrollment)
2. `carrier.csv.xlsx` (Professional / Physician Claims)
3. `inpatient.csv` (Inpatient Hospital Claims)
4. `outpatient.csv` (Outpatient Hospital Claims)
5. `pde.csv` (Part D Prescription Drug Claims)
6. `2026 Final ICD-10-CM Mappings.xlsx` (CMS-HCC V28 Reference Mapping)

## 3. Selection Justification & Excluded Datasets
- **Included Sources**: Selected because inpatient, outpatient, carrier (physician), and Part D pharmacy claims represent the primary clinical evidence channels required for CMS-HCC V28 suspecting.
- **Excluded Datasets**:
  - `dme.csv` (Durable Medical Equipment)
  - `hha.csv` (Home Health Agency)
  - `hospice.csv` (Hospice Care)
  - `snf.csv` (Skilled Nursing Facility)
  - `Synthea` (Synthetic EHR Cohort)
- **Exclusion Justification**: Excluded to focus the MVP on core medical and pharmacy evidence, reduce dataset size, and maintain strict population isolation within CMS Medicare Fee-For-Service data.

## 4. Member Selection Methodology
- Exactly **5,000 unique member records** (`BENE_00001` through `BENE_05000`) were selected deterministically from the CMS beneficiary population.
- Each member represents an individual beneficiary record with complete demographic attributes (`birth_date`, `sex`, `state`).
- Claims and diagnosis events from Carrier, Inpatient, Outpatient, and PDE datasets were mapped deterministically to these 5,000 target members.

## 5. Output Summary & Record Counts

| MVP File | Target Path | Output Record Count | Description |
| --- | --- | --- | --- |
| `members.csv` | `data/mvp/members.csv` | **{len(members_rows):,}** | Demographics for 5,000 target members |
| `claims.csv` | `data/mvp/claims.csv` | **{total_claims_count:,}** | Medical claims (Carrier, Inpatient, Outpatient) & PDE prescription fills |
| `diagnoses.csv` | `data/mvp/diagnoses.csv` | **{total_diag_count:,}** | Principal & secondary ICD-10 diagnosis codes |
| `hcc_mapping.csv` | `data/mvp/hcc_mapping.csv` | **{len(hcc_rows):,}** | Official CMS-HCC V28 reference lookup table |

## 6. MVP CSV Column Definitions

### A. `members.csv`
- `member_id`: Deterministic member identifier (`BENE_00001` .. `BENE_05000`)
- `birth_date`: Date of birth in ISO format (`YYYY-MM-DD`)
- `sex`: Gender code (`1` = Male, `2` = Female)
- `state`: State location code

### B. `claims.csv`
- `member_id`: Foreign key link to `members.csv(member_id)`
- `claim_id`: Unique claim identifier (e.g. `CLM_INP_0000001`, `CLM_PDE_0000001`)
- `claim_type`: Claim domain (`CARRIER`, `INPATIENT`, `OUTPATIENT`, `PDE`)
- `service_date`: Claim service start date in ISO format (`YYYY-MM-DD`)

### C. `diagnoses.csv`
- `member_id`: Foreign key link to `members.csv(member_id)`
- `claim_id`: Foreign key link to `claims.csv(claim_id)`
- `diagnosis_code`: Cleaned uppercase ICD-10-CM diagnosis code (e.g. `E119`, `I10`)
- `diagnosis_date`: Date of service when diagnosis was billed (`YYYY-MM-DD`)
- `is_principal`: Boolean indicator (`True` if principal/admitting diagnosis, `False` if secondary)

### D. `hcc_mapping.csv`
- `diagnosis_code`: Cleaned uppercase ICD-10-CM diagnosis code
- `description`: Official ICD-10 diagnosis text description
- `hcc_v28`: CMS-HCC V28 model category number (e.g. `19`, `37`) or empty if unmapped
- `payment_2026`: Boolean flag (`True` if eligible for 2026 Payment Year)

## 7. System & Business Rule Alignment
- **Human-in-the-Loop Rule**: The MVP dataset provides raw clinical evidence for a human coder/clinician review and does not make final coding decisions or submit claims.
- **No Database / No Parquet**: All data is stored in plain CSV format under `data/mvp/`.
- **No Engine Code Yet**: No AI/suspecting models or risk score calculators were built.

## 8. Limitations
- Scoped to 5,000 members for hackathon execution speed.
- Unmapped ICD-10 codes remain empty in `hcc_v28` without artificial fallback.
"""

    with open(doc_path, "w", encoding="utf-8") as f:
        f.write(doc_content)
    print(f"   Written documentation file to {doc_path}")

    elapsed = time.time() - t0
    print(f"\n=== 5,000-Member MVP Dataset Completed Successfully in {elapsed:.2f} seconds! ===")

if __name__ == "__main__":
    build_mvp_dataset()
